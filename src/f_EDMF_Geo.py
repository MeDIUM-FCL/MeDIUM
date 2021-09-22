def EDMFanalysis(numberofsensor,numberofmodelinstance,meanuncertaintydata,stduncertaintydata,ID_to_be_excluded,finalpathprediction,finalpathmeas,sidakvalue,uncmultiplier,indextext):
    """
        Perform EDMF analysis with all uploaded files and intermediate results

        Developed by :   WANG Ze-Zhou (ETH Singapore)
        Contact :        e0054291@u.nus.edu
        Date:            August 03, 2021

        INPUTS:
            numberofsensor : number of sensors involved in the studied problem.
            numberofmodelinstance : number of initial model instances involved in the studied problem.
            ID_to_be_excluded: index of measurements to be excluded for EDMF analysis.
            finalpathprediction : directory of the Prediction(Geo).xlsx.
            finalpathmeas: directory of the Measurements(Geo).xlsx.
            sidakvalue: target reliability of identification
            uncmultiplier : multiplier for uncertainty to assess sensitivity of results to uncertainty definitions
            indextext: index of measurements to be excluded for EDMF analysis in the original format.
        OUTPUTS:
            prediction : predictions made with candidate models.
            measurement : measurement data.
            finalresults: logic results of initial model instances: 0 represents falsified model; 1 represents candidate model
            IMS: initial model instances
            CMS: candidate models
            CMSID: location reference of candidate models
        NOTE:
            These inputs will be automatically read from other functions. No maunal actions are needed.
        """

    import numpy
    import os
    import xlrd
    from openpyxl import load_workbook
    import pandas
    from scipy.stats import norm


    #########for validation
    fullset = numpy.arange(numberofsensor)+1
    ID_to_be_included_ = set(ID_to_be_excluded) ^ set(fullset)
    ID_to_be_included = numpy.zeros((len(ID_to_be_included_), 1))
    for i in range(0, len(ID_to_be_included_)):
        ID_to_be_included[i][0] = (list(ID_to_be_included_)[i])
    ID_to_be_included = ID_to_be_included.astype(numpy.int64)-1

    #####load uncertainty excel
    excelfileuncertainty1_ = meanuncertaintydata
    excelfileuncertainty2_ = stduncertaintydata

    excelfileprediction = pandas.read_excel(finalpathprediction,sheet_name = 'Prediction', engine = 'openpyxl')

    excelfilemeas = pandas.read_excel(finalpathmeas,header = None,sheet_name = 'Measurement', engine = 'openpyxl')

    ##########read uncertainty
    excelsheetuncertainty1 = excelfileuncertainty1_[0:numberofmodelinstance, 0:numberofsensor]
    ucombinemean = excelsheetuncertainty1

    excelsheetuncertainty2 = excelfileuncertainty2_[0:numberofmodelinstance, 0:numberofsensor]
    ucombinesd = excelsheetuncertainty2


    ###########calculate sidak
    lowerbound = numpy.zeros((numberofmodelinstance, numberofsensor))
    upperbound = numpy.zeros((numberofmodelinstance, numberofsensor))

    sidak = sidakvalue ** (1 / len(ID_to_be_included))

    for i in range(0,numberofmodelinstance):
        for s in range(0,numberofsensor):
            bound1_ = norm.ppf((1-sidak)/2,loc = ucombinemean[i][s]*uncmultiplier,scale = ucombinesd[i][s])
            bound2_ = norm.ppf(1-(1-sidak)/2,loc = ucombinemean[i][s]*uncmultiplier,scale = ucombinesd[i][s])
            lowerbound[i][s] = bound1_
            upperbound[i][s] = bound2_


    #######falsification
    #######load predictions and measurements from other functions or excel
    excelsheetprediction = excelfileprediction.to_numpy()
    prediction = numpy.zeros((numberofmodelinstance, numberofsensor))
    for i in range(0,numberofmodelinstance):
        for s in range(0,numberofsensor):
            excelsheetpredictionvalue_= excelsheetprediction[i,s+2]
            prediction[i][s] = excelsheetpredictionvalue_

    excelsheetmeas = excelfilemeas.to_numpy()
    measurement = numpy.zeros((numberofsensor,1))
    for s in range(0,numberofsensor):
        excelsheetmeasvalue_= excelsheetmeas[0,s]
        measurement[s][0] = excelsheetmeasvalue_

    falsification = numpy.zeros((numberofmodelinstance, numberofsensor))

    for i in range(0,numberofmodelinstance):
        for s in range(0,numberofsensor):
            residual = prediction[i][s] - measurement[s][0]
            falsification_ = (residual<upperbound[i][s] and residual>lowerbound[i][s])
            if falsification_==True:
                falsification[i][s] = 1
            else:
                falsification[i][s] = 0

    ##########for validation
    falsificationfinal = numpy.zeros((numberofmodelinstance,len(ID_to_be_included)))
    for i in range(0,len(ID_to_be_included)):
        for s in range(0,numberofmodelinstance):
            falsificationfinal[s][i] = falsification[s,int(ID_to_be_included[i][0])]

    final = sum(numpy.transpose(falsificationfinal))
    is_candidate = numpy.zeros((numberofmodelinstance,1))

    for i in range(0,numberofmodelinstance):
        if final[i] == float(len(ID_to_be_included)):
            is_candidate[i][0] = 1
        else:
            is_candidate[i][0] = 0

    ###write to excel
    finalpathCMS = finalpathprediction
    savefile = load_workbook(finalpathCMS)
    savesheets = savefile.sheetnames
    sheetCMS = savefile[savesheets[1]]

    for s in range(0,numberofmodelinstance):
        sheetCMS.cell(row = s+3, column = 5).value = is_candidate[s][0]

    savefile.save(finalpathCMS)

    ###########load IMS
    excelsheetIMS = excelfileprediction.to_numpy()[0:numberofmodelinstance, 0:2]
    ims = numpy.zeros((numberofmodelinstance, 2))
    for i in range(0,numberofmodelinstance):
        for s in range(0,2):
            excelsheetIMSvalue_= excelsheetIMS[i,s]
            ims[i][s] = excelsheetIMSvalue_

    ########CMS
    cmsID = numpy.where(is_candidate == 1)
    cms = ims[cmsID[0], :]

    ########write to excel
    from f_resultsexcel_Geo import f_resultsexcel_Geo
    f_resultsexcel_Geo(ims, prediction, is_candidate,uncmultiplier,sidakvalue, indextext)

    return prediction,measurement,is_candidate,ims,cms,cmsID,sidakvalue,indextext,uncmultiplier